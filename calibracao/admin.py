from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django import forms
from django.contrib import admin, messages
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.html import format_html, format_html_join
from django.utils.safestring import mark_safe
from django.urls import path, reverse
from openpyxl import Workbook, load_workbook

from clientes.models import Cliente
from qualidade.models import Documento

from .models import (
    Instrumento,
    InstrumentoTecnico,
    OrdemServico,
    Padrao,
    Periodicidade,
    Calibracao,
    CalibracaoAnexo,
    CalibracaoTurbidez,
    TurbidezPadraoUtilizado,
    TurbidezVerificacaoPonto,
    TurbidezCalibracaoPonto,
    TurbidezIncertezaPonto,
    ResponsavelCertificado,
    CalibracaoColorimetro,
    ColorimetroPadraoUtilizado,
    ColorimetroVerificacaoPonto,
    ColorimetroCalibracaoPonto,
    ColorimetroIncertezaPonto,
    CalibracaoPressao,
    PressaoPadraoUtilizado,
    PressaoCalibracaoPonto,
    PressaoIncertezaPonto,
    CalibracaoCondutividade,
)
from .ph_models import (
    CalibracaoPH,
    CalibracaoPHPadraoUtilizado,
    CalibracaoPHPonto,
    CalibracaoPHIncertezaPonto,
    _join_address,
)


# =========================
# INLINES
# =========================

class PeriodicidadeInline(admin.TabularInline):
    model = Periodicidade
    extra = 1


class InstrumentoTecnicoInline(admin.StackedInline):
    model = InstrumentoTecnico
    extra = 0
    fields = (
        "faixa_medicao",
        "capacidade_total",
        "menor_resolucao",
        "unidade",
        "classe",
        "observacoes",
    )


class InstrumentoAdminForm(forms.ModelForm):
    class Meta:
        model = Instrumento
        fields = "__all__"

    class Media:
        js = ("js/admin_instrumento_duplicate_alert.js",)


class InstrumentoImportForm(forms.Form):
    arquivo = forms.FileField(
        label="Planilha Excel",
        help_text="Envie um arquivo .xlsx seguindo o modelo disponibilizado.",
    )


# =========================
# RESPONSÁVEIS
# =========================

@admin.register(ResponsavelCertificado)
class ResponsavelCertificadoAdmin(admin.ModelAdmin):
    list_display = ("nome", "cargo", "ativo")
    list_filter = ("ativo",)
    search_fields = ("nome", "cargo")
    ordering = ("nome",)


# =========================
# INSTRUMENTO
# =========================

@admin.register(Instrumento)
class InstrumentoAdmin(admin.ModelAdmin):
    form = InstrumentoAdminForm
    change_list_template = "admin/calibracao/instrumento/change_list.html"

    list_display = (
        "codigo",
        "descricao",
        "marca",
        "modelo",
        "numero_serie",
        "tag",
        "cliente",
        "status",
        "ver_certificados_padroes",
    )

    list_filter = (
        "status",
        "cliente",
    )

    search_fields = (
        "codigo",
        "descricao",
        "marca",
        "modelo",
        "numero_serie",
        "tag",
        "cliente__razao_social",
    )

    autocomplete_fields = (
        "cliente",
        "padroes",
    )

    readonly_fields = (
        "ver_certificados_padroes",
        )

    ordering = ("descricao",)

    fieldsets = (
        ("Identificação", {
            "fields": (
                "codigo",
                "descricao",
                "cliente",
                "status",
            )
        }),

        ("Dados Técnicos", {
            "fields": (
                "marca",
                "modelo",
                "numero_serie",
                "tag",
            )
        }),

        ("Localização", {
            "fields": (
                "local_instalacao",
            )
        }),

        ("Metrologia", {
            "fields": (
                "metodo_calibracao",
                "padroes",
                "ver_certificados_padroes",
            )
        }),

        ("Controle", {
            "fields": (
                "proxima_calibracao",
                "ativo",
                "nome_anexo",
                "anexo",
            )
        }),
    )

    inlines = [
        InstrumentoTecnicoInline,
        PeriodicidadeInline,
    ]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "importar-excel/",
                self.admin_site.admin_view(self.importar_excel_view),
                name="calibracao_instrumento_importar_excel",
            ),
            path(
                "modelo-importacao/",
                self.admin_site.admin_view(self.download_modelo_importacao_view),
                name="calibracao_instrumento_modelo_importacao",
            ),
            path(
                "exportar-excel/",
                self.admin_site.admin_view(self.exportar_excel_view),
                name="calibracao_instrumento_exportar_excel",
            ),
            path(
                "check-duplicate/",
                self.admin_site.admin_view(self.check_duplicate_view),
                name="calibracao_instrumento_check_duplicate",
            )
        ]
        return custom_urls + urls

    def check_duplicate_view(self, request):
        cliente_id = request.GET.get("cliente_id")
        numero_serie = (request.GET.get("numero_serie") or "").strip()
        codigo = ""
        tag = (request.GET.get("tag") or "").strip()
        object_id = request.GET.get("object_id")

        if not cliente_id:
            return JsonResponse({"duplicate": False})

        if numero_serie and tag:
            qs_serie = Instrumento.objects.filter(
                cliente_id=cliente_id,
                numero_serie__iexact=numero_serie,
                tag__iexact=tag,
            )
            if object_id:
                qs_serie = qs_serie.exclude(pk=object_id)
            if qs_serie.exists():
                return JsonResponse(
                    {
                        "duplicate": True,
                        "message": "Já existe instrumento cadastrado para este cliente com o mesmo número de série e a mesma TAG.",
                    }
                )

        if codigo:
            qs_codigo = Instrumento.objects.filter(codigo__iexact=codigo)
            if object_id:
                qs_codigo = qs_codigo.exclude(pk=object_id)
            if qs_codigo.exists():
                return JsonResponse(
                    {
                        "duplicate": True,
                        "message": "Já existe instrumento cadastrado com a mesma TAG/código.",
                    }
                )

        return JsonResponse({"duplicate": False})

    def importar_excel_view(self, request):
        if not self.has_add_permission(request):
            return redirect("admin:calibracao_instrumento_changelist")

        form = InstrumentoImportForm(request.POST or None, request.FILES or None)
        if request.method == "POST" and form.is_valid():
            try:
                resultado = self._processar_planilha_importacao(form.cleaned_data["arquivo"])
            except ValueError as exc:
                self.message_user(request, str(exc), level=messages.ERROR)
            else:
                if resultado["criados"]:
                    self.message_user(
                        request,
                        f'{resultado["criados"]} instrumento(s) importado(s) com sucesso.',
                        level=messages.SUCCESS,
                    )
                if resultado["ignorados"]:
                    self.message_user(
                        request,
                        f'{resultado["ignorados"]} linha(s) ignorada(s) por estarem vazias.',
                        level=messages.WARNING,
                    )
                if resultado.get("sem_tag"):
                    self.message_user(
                        request,
                        f'{resultado["sem_tag"]} linha(s) importada(s) sem TAG. Isso e permitido, mas reduz a rastreabilidade operacional.',
                        level=messages.WARNING,
                    )
                if resultado["erros"]:
                    resumo = " | ".join(resultado["erros"][:10])
                    if len(resultado["erros"]) > 10:
                        resumo += f' | ... e mais {len(resultado["erros"]) - 10} erro(s).'
                    self.message_user(request, resumo, level=messages.ERROR)
                if not any(resultado.values()):
                    self.message_user(
                        request,
                        "Nenhuma linha válida foi encontrada na planilha.",
                        level=messages.WARNING,
                    )
                return redirect("admin:calibracao_instrumento_changelist")

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Importar instrumentos por Excel",
            "subtitle": "Cadastro em lote de instrumentos",
            "form": form,
            "download_model_url": reverse("admin:calibracao_instrumento_modelo_importacao"),
        }
        return render(request, "admin/calibracao/instrumento/importar_excel.html", context)

    def download_modelo_importacao_view(self, request):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Instrumentos"
        headers = [
            "cliente_codigo",
            "cliente_cnpj",
            "cliente_razao_social",
            "codigo",
            "descricao",
            "marca",
            "modelo",
            "numero_serie",
            "tag",
            "local_instalacao",
            "status",
            "proxima_calibracao",
            "ativo",
            "faixa_medicao",
            "capacidade_total",
            "menor_resolucao",
            "unidade",
            "classe",
            "observacoes",
        ]
        worksheet.append(headers)
        worksheet.append(
            [
                "CL-0001",
                "",
                "",
                "EQ-0001",
                "Manômetro analógico",
                "Wika",
                "232.50",
                "SN-12345",
                "PT-101",
                "Linha 1",
                "ativo",
                "2026-12-31",
                "sim",
                "0 a 10 bar",
                "10 bar",
                "0.1000",
                "bar",
                "A",
                "Cadastro inicial via importação",
            ]
        )
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 15), 28)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="modelo_importacao_instrumentos.xlsx"'
        return response

    def exportar_excel_view(self, request):
        changelist = self.get_changelist_instance(request)
        queryset = changelist.get_queryset(request).select_related("cliente").prefetch_related("padroes")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Instrumentos"
        headers = [
            "cliente_codigo",
            "cliente_razao_social",
            "cliente_cnpj",
            "codigo",
            "descricao",
            "marca",
            "modelo",
            "numero_serie",
            "tag",
            "local_instalacao",
            "status",
            "proxima_calibracao",
            "ativo",
            "faixa_medicao",
            "capacidade_total",
            "menor_resolucao",
            "unidade",
            "classe",
            "observacoes",
            "metodo_calibracao",
            "padroes",
        ]
        worksheet.append(headers)

        for instrumento in queryset:
            try:
                tecnico = instrumento.tecnico
            except InstrumentoTecnico.DoesNotExist:
                tecnico = None

            worksheet.append(
                [
                    getattr(instrumento.cliente, "codigo", ""),
                    getattr(instrumento.cliente, "razao_social", ""),
                    getattr(instrumento.cliente, "cnpj", ""),
                    instrumento.codigo,
                    instrumento.descricao,
                    instrumento.marca,
                    instrumento.modelo,
                    instrumento.numero_serie,
                    instrumento.tag,
                    instrumento.local_instalacao,
                    instrumento.status,
                    instrumento.proxima_calibracao.isoformat() if instrumento.proxima_calibracao else "",
                    "sim" if instrumento.ativo else "nao",
                    getattr(tecnico, "faixa_medicao", "") if tecnico else "",
                    getattr(tecnico, "capacidade_total", "") if tecnico else "",
                    str(getattr(tecnico, "menor_resolucao", "") or "") if tecnico else "",
                    getattr(tecnico, "unidade", "") if tecnico else "",
                    getattr(tecnico, "classe", "") if tecnico else "",
                    getattr(tecnico, "observacoes", "") if tecnico else "",
                    getattr(instrumento.metodo_calibracao, "codigo", "") if instrumento.metodo_calibracao else "",
                    ", ".join(instrumento.padroes.values_list("codigo", flat=True)),
                ]
            )

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 15), 32)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="instrumentos_exportados.xlsx"'
        return response

    def _processar_planilha_importacao(self, arquivo):
        nome = (getattr(arquivo, "name", "") or "").lower()
        if not nome.endswith(".xlsx"):
            raise ValueError("Formato inválido. Envie uma planilha .xlsx.")

        try:
            workbook = load_workbook(arquivo, data_only=True)
        except Exception as exc:
            raise ValueError(f"Não foi possível ler a planilha enviada: {exc}") from exc

        worksheet = workbook.active
        linhas = list(worksheet.iter_rows(values_only=True))
        if not linhas:
            raise ValueError("A planilha está vazia.")

        headers = [self._normalizar_cabecalho_excel(valor) for valor in linhas[0]]
        obrigatorios = {"codigo", "descricao"}
        ausentes = sorted(campo for campo in obrigatorios if campo not in headers)
        if ausentes:
            raise ValueError(
                "A planilha não contém as colunas obrigatórias: " + ", ".join(ausentes) + "."
            )

        criados = 0
        ignorados = 0
        sem_tag = 0
        erros = []
        for indice, valores in enumerate(linhas[1:], start=2):
            registro = {
                headers[posicao]: valores[posicao] if posicao < len(valores) else None
                for posicao in range(len(headers))
                if headers[posicao]
            }
            if self._linha_excel_vazia(registro):
                ignorados += 1
                continue
            if not self._texto_excel(registro.get("tag")):
                sem_tag += 1
            try:
                with transaction.atomic():
                    self._criar_instrumento_importado(registro)
                criados += 1
            except Exception as exc:
                erros.append(f"Linha {indice}: {exc}")

        return {"criados": criados, "ignorados": ignorados, "sem_tag": sem_tag, "erros": erros}

    def _criar_instrumento_importado(self, registro):
        cliente = self._resolver_cliente_importacao(registro)
        codigo = self._texto_excel(registro.get("codigo"))
        descricao = self._texto_excel(registro.get("descricao"))
        numero_serie = self._texto_excel(registro.get("numero_serie"))
        tag = self._texto_excel(registro.get("tag"))

        if not codigo:
            raise ValueError("campo 'codigo' é obrigatório.")
        if not descricao:
            raise ValueError("campo 'descricao' é obrigatório.")
        if not cliente:
            raise ValueError(
                "cliente não encontrado. Informe 'cliente_codigo', 'cliente_cnpj' ou 'cliente_razao_social'."
            )
        if Instrumento.objects.filter(codigo__iexact=codigo).exists():
            raise ValueError(f"já existe instrumento cadastrado com o código '{codigo}'.")
        if tag and Instrumento.objects.filter(
            cliente=cliente,
            tag__iexact=tag,
        ).exists():
            raise ValueError(
                "já existe instrumento para este cliente com o mesmo número de série e a mesma TAG."
            )

        instrumento = Instrumento(
            cliente=cliente,
            codigo=codigo,
            descricao=descricao,
            marca=self._texto_excel(registro.get("marca")),
            modelo=self._texto_excel(registro.get("modelo")),
            numero_serie=numero_serie or None,
            tag=tag,
            local_instalacao=self._texto_excel(registro.get("local_instalacao")),
            status=self._normalizar_status_importacao(registro.get("status")),
            proxima_calibracao=self._normalizar_data_excel(registro.get("proxima_calibracao")),
            ativo=self._normalizar_booleano_excel(registro.get("ativo"), default=True),
        )
        instrumento.full_clean()
        instrumento.save()

        tecnico_campos = {
            "faixa_medicao": self._texto_excel(registro.get("faixa_medicao")),
            "capacidade_total": self._texto_excel(registro.get("capacidade_total")),
            "menor_resolucao": self._normalizar_decimal_excel(registro.get("menor_resolucao")),
            "unidade": self._texto_excel(registro.get("unidade")),
            "classe": self._texto_excel(registro.get("classe")),
            "observacoes": self._texto_excel(registro.get("observacoes")),
        }
        if any(valor not in (None, "") for valor in tecnico_campos.values()):
            InstrumentoTecnico.objects.update_or_create(
                instrumento=instrumento,
                defaults=tecnico_campos,
            )

    def _resolver_cliente_importacao(self, registro):
        cliente_codigo = self._texto_excel(registro.get("cliente_codigo"))
        cliente_cnpj = self._texto_excel(registro.get("cliente_cnpj"))
        cliente_razao = self._texto_excel(registro.get("cliente_razao_social"))

        if cliente_codigo:
            cliente = Cliente.objects.filter(codigo__iexact=cliente_codigo).first()
            if cliente:
                return cliente
        if cliente_cnpj:
            cliente = Cliente.objects.filter(cnpj__iexact=cliente_cnpj).first()
            if cliente:
                return cliente
        if cliente_razao:
            cliente = Cliente.objects.filter(razao_social__iexact=cliente_razao).first()
            if cliente:
                return cliente
        return None

    @staticmethod
    def _texto_excel(valor):
        if valor is None:
            return ""
        if isinstance(valor, str):
            return valor.strip()
        return str(valor).strip()

    @staticmethod
    def _normalizar_cabecalho_excel(valor):
        texto = InstrumentoAdmin._texto_excel(valor).lower()
        substituicoes = {
            "á": "a",
            "à": "a",
            "ã": "a",
            "â": "a",
            "é": "e",
            "ê": "e",
            "í": "i",
            "ó": "o",
            "ô": "o",
            "õ": "o",
            "ú": "u",
            "ç": "c",
        }
        for origem, destino in substituicoes.items():
            texto = texto.replace(origem, destino)
        return texto.replace(" ", "_")

    @staticmethod
    def _linha_excel_vazia(registro):
        return not any(valor not in (None, "") for valor in registro.values())

    @staticmethod
    def _normalizar_status_importacao(valor):
        texto = InstrumentoAdmin._texto_excel(valor).lower()
        if texto in {"", "ativo"}:
            return "ativo"
        if texto == "inativo":
            return "inativo"
        if texto in {"manutencao", "manutenção", "em manutencao", "em manutenção"}:
            return "manutencao"
        raise ValueError(f"status inválido: '{valor}'. Use ativo, inativo ou manutencao.")

    @staticmethod
    def _normalizar_booleano_excel(valor, default=True):
        texto = InstrumentoAdmin._texto_excel(valor).lower()
        if texto == "":
            return default
        if texto in {"1", "sim", "s", "true", "verdadeiro", "ativo"}:
            return True
        if texto in {"0", "nao", "não", "n", "false", "falso", "inativo"}:
            return False
        raise ValueError(f"valor booleano inválido: '{valor}'. Use sim/não.")

    @staticmethod
    def _normalizar_data_excel(valor):
        if valor in (None, ""):
            return None
        if isinstance(valor, datetime):
            return valor.date()
        if hasattr(valor, "year") and hasattr(valor, "month") and hasattr(valor, "day"):
            return valor
        texto = InstrumentoAdmin._texto_excel(valor)
        for formato in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(texto, formato).date()
            except ValueError:
                continue
        raise ValueError(f"data inválida: '{valor}'. Use YYYY-MM-DD ou DD/MM/AAAA.")

    @staticmethod
    def _normalizar_decimal_excel(valor):
        if valor in (None, ""):
            return None
        if isinstance(valor, Decimal):
            return valor
        try:
            return Decimal(str(valor).replace(",", "."))
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise ValueError(f"valor decimal inválido: '{valor}'.") from exc

    def ver_certificados_padroes(self, obj):
        links = [
            (p.certificado.url, p.codigo)
            for p in obj.padroes.all()
            if p.certificado
        ]
        if not links:
            return "?"
        return format_html_join(
            mark_safe("<br>"),
            '<a href="{}" target="_blank">{}</a>',
            links,
        )

    ver_certificados_padroes.short_description = "Certificados dos Padrões"


# =========================
# CALIBRAÇÃO
# =========================
class CalibracaoAnexoInline(admin.TabularInline):
    model = CalibracaoAnexo
    extra = 1


class TurbidezPadraoInline(admin.TabularInline):
    model = TurbidezPadraoUtilizado
    extra = 1
    autocomplete_fields = ("padrao",)
    fields = (
        "tipo",
        "ordem",
        "padrao",
        "codigo",
        "descricao",
        "numero_certificado",
        "laboratorio_emitente",
        "data_calibracao",
        "validade",
        "resolucao",
        "incerteza",
        "fator_k",
        "graus_liberdade",
        "unidade",
        "valor_nominal",
    )
    verbose_name = "Padrão"
    verbose_name_plural = "Padrões"


class TurbidezVerificacaoInline(admin.TabularInline):
    model = TurbidezVerificacaoPonto
    extra = 5
    fields = ("ordem", "valor_padrao", "leitura", "erro", "criterio", "resultado")
    readonly_fields = ("erro",)
    verbose_name = "Ponto de verificação"
    verbose_name_plural = "Verificação"


class TurbidezCalibracaoInline(admin.TabularInline):
    model = TurbidezCalibracaoPonto
    extra = 9
    fields = (
        "ordem",
        "valor_referencia",
        "leitura_1",
        "leitura_2",
        "leitura_3",
        "media",
        "erro",
        "ema",
        "criterio",
    )
    readonly_fields = ("media", "erro", "ema")
    verbose_name = "Ponto de calibração"
    verbose_name_plural = "Calibração"


class TurbidezIncertezaInline(admin.TabularInline):
    model = TurbidezIncertezaPonto
    extra = 9
    fields = (
        "ordem",
        "repetibilidade",
        "resolucao_instrumento",
        "incerteza_padrao",
        "resolucao_padrao",
        "incerteza_curva",
        "incerteza_turbidimetro",
        "fator_k",
        "graus_liberdade",
        "incerteza_padrao_combinada",
        "incerteza_expandida",
    )
    readonly_fields = ("incerteza_padrao_combinada", "incerteza_expandida")
    verbose_name = "Ponto de incerteza"
    verbose_name_plural = "Incerteza"

    def get_readonly_fields(self, request, obj=None):
        campos_calculados = ("incerteza_padrao_combinada", "incerteza_expandida")
        if obj and not request.user.is_superuser:
            return (
                "ordem",
                "repetibilidade",
                "resolucao_instrumento",
                "incerteza_padrao",
                "resolucao_padrao",
                "incerteza_curva",
                "incerteza_turbidimetro",
                "fator_k",
                "graus_liberdade",
                *campos_calculados,
            )
        return campos_calculados


class ColorimetroPadraoInline(admin.TabularInline):
    model = ColorimetroPadraoUtilizado
    extra = 1
    autocomplete_fields = ("padrao",)
    fields = (
        "tipo",
        "ordem",
        "padrao",
        "codigo",
        "descricao",
        "numero_certificado",
        "laboratorio_emitente",
        "data_calibracao",
        "validade",
        "resolucao",
        "incerteza",
        "fator_k",
        "graus_liberdade",
        "unidade",
        "valor_nominal",
    )
    verbose_name = "Padrão"
    verbose_name_plural = "Padrões"


class ColorimetroVerificacaoInline(admin.TabularInline):
    model = ColorimetroVerificacaoPonto
    extra = 5
    fields = ("ordem", "valor_padrao", "leitura", "erro", "criterio", "criterio_origem", "resultado")
    readonly_fields = ("erro",)
    verbose_name = "Ponto de verificação"
    verbose_name_plural = "Verificação"


class ColorimetroCalibracaoInline(admin.TabularInline):
    model = ColorimetroCalibracaoPonto
    extra = 9
    fields = (
        "ordem",
        "valor_referencia",
        "leitura_1",
        "leitura_2",
        "leitura_3",
        "media",
        "erro",
        "ema",
        "criterio",
        "criterio_origem",
    )
    readonly_fields = ("media", "erro", "ema")
    verbose_name = "Ponto de calibração"
    verbose_name_plural = "Calibração"


class ColorimetroIncertezaInline(admin.TabularInline):
    model = ColorimetroIncertezaPonto
    extra = 9
    fields = (
        "ordem",
        "repetibilidade",
        "resolucao_instrumento",
        "incerteza_padrao",
        "resolucao_padrao",
        "incerteza_curva",
        "fator_k",
        "graus_liberdade",
        "incerteza_padrao_combinada",
        "incerteza_expandida",
    )
    readonly_fields = ("incerteza_padrao_combinada", "incerteza_expandida")
    verbose_name = "Ponto de incerteza"
    verbose_name_plural = "Incerteza"


class PressaoPadraoInlineForm(forms.ModelForm):
    class Meta:
        model = PressaoPadraoUtilizado
        fields = "__all__"

    def clean(self):
        cleaned_data = super().clean()
        if cleaned_data.get("DELETE"):
            return cleaned_data

        padrao = cleaned_data.get("padrao")

        possui_dados = any(
            cleaned_data.get(campo) not in (None, "")
            for campo in (
                "padrao",
                "codigo",
                "descricao",
                "numero_certificado",
                "laboratorio_emitente",
                "data_calibracao",
                "validade",
                "incerteza",
                "resolucao",
            )
        )
        if not possui_dados:
            return cleaned_data

        campos_criticos = {
            "codigo": "Identificação do padrão",
            "descricao": "Descrição do padrão",
            "numero_certificado": "Número do certificado",
            "laboratorio_emitente": "Laboratório",
            "data_calibracao": "Data da calibração",
            "validade": "Data da validade",
        }
        for campo, rotulo in campos_criticos.items():
            valor = cleaned_data.get(campo)
            if valor in (None, "") and padrao is not None:
                if campo == "validade":
                    valor = getattr(padrao, "vencimento", None)
                else:
                    valor = getattr(padrao, campo, None)
            if valor in (None, ""):
                self.add_error(campo, f"{rotulo} é obrigatória para rastreabilidade metrológica.")

        return cleaned_data


class PressaoPadraoInline(admin.TabularInline):
    model = PressaoPadraoUtilizado
    form = PressaoPadraoInlineForm
    extra = 3
    fields = (
        "tipo",
        "ordem",
        "padrao",
        "codigo",
        "descricao",
        "tipo_padrao",
        "numero_certificado",
        "laboratorio_emitente",
        "data_calibracao",
        "validade",
        "resolucao",
        "incerteza",
        "unidade",
        "status_validade",
    )
    autocomplete_fields = ("padrao",)
    readonly_fields = ("status_validade",)
    verbose_name = "Padrão utilizado"
    verbose_name_plural = "Padrões"


class PressaoCalibracaoInlineForm(forms.ModelForm):
    class Meta:
        model = PressaoCalibracaoPonto
        fields = "__all__"
        labels = {
            "valor_referencia": "Padrão em",
            "valor_referencia_convertido": "Padrão convertido",
            "leitura_1": "Crescente 1ª série",
            "leitura_2": "Crescente 2ª série",
            "leitura_3": "Decrescente 1ª série",
            "leitura_4": "Decrescente 2ª série",
            "media": "Média",
            "desvio_padrao": "Desvio padrão",
            "erro": "Erro do instrumento",
            "erro_percentual": "Erro (%)",
            "ema": "EMA",
            "criterio": "Critério de aceitação",
            "criterio_origem": "Origem da tolerância",
            "criterio_referencia": "Referência do critério",
            "resultado": "Resultado",
        }

    def clean(self):
        cleaned_data = super().clean()
        if self.instance and self.instance.pk and not cleaned_data.get("DELETE"):
            for campo in ("leitura_1", "leitura_2", "leitura_3", "leitura_4"):
                if cleaned_data.get(campo) is None and getattr(self.instance, campo) is not None:
                    cleaned_data[campo] = getattr(self.instance, campo)
        return cleaned_data


class CalibracaoPressaoAdminForm(forms.ModelForm):
    class Meta:
        model = CalibracaoPressao
        fields = "__all__"

    def clean(self):
        cleaned_data = super().clean()
        campos_criticos = {
            "temperatura_inicial": "Temperatura inicial",
            "temperatura_final": "Temperatura final",
            "umidade_inicial": "Umidade inicial",
            "umidade_final": "Umidade final",
        }
        for campo, rotulo in campos_criticos.items():
            if cleaned_data.get(campo) in (None, ""):
                self.add_error(campo, f"{rotulo} é obrigatória para emissão do certificado.")

        if not cleaned_data.get("procedimento_documento") and not (cleaned_data.get("procedimento_numero") or "").strip():
            self.add_error(
                "procedimento_documento",
                "Informe o procedimento aplicável para emissão do certificado.",
            )

        return cleaned_data


class PressaoCalibracaoInline(admin.TabularInline):
    model = PressaoCalibracaoPonto
    form = PressaoCalibracaoInlineForm
    extra = 10
    fields = (
        "ordem",
        "valor_referencia",
        "valor_referencia_convertido",
        "leitura_1",
        "leitura_2",
        "leitura_3",
        "leitura_4",
        "media",
        "desvio_padrao",
        "erro",
        "erro_percentual",
        "ema",
        "criterio",
        "criterio_origem",
        "criterio_referencia",
        "resultado",
    )
    readonly_fields = (
        "valor_referencia_convertido",
        "media",
        "desvio_padrao",
        "erro",
        "erro_percentual",
        "ema",
    )
    verbose_name = "Ponto de calibração"
    verbose_name_plural = "Calibração"


class PressaoIncertezaInline(admin.TabularInline):
    model = PressaoIncertezaPonto
    extra = 10
    fields = (
        "ordem",
        "repetibilidade",
        "resolucao_instrumento",
        "resolucao_padrao",
        "incerteza_padrao",
        "incerteza_curva",
        "fator_k",
        "graus_liberdade",
        "incerteza_padrao_combinada",
        "incerteza_expandida",
    )
    readonly_fields = ("incerteza_padrao_combinada", "incerteza_expandida")
    verbose_name = "Ponto de incerteza"
    verbose_name_plural = "Incerteza"

@admin.register(Calibracao)
class CalibracaoAdmin(admin.ModelAdmin):

    list_display = (
        "instrumento",
        "data_calibracao",
        "validade",
        "certificado_numero",
        "resultado",
        "local_calibracao",
        "empresa_emissora",
        "ver_anexo",
    )

    list_filter = (
        "resultado",
        "empresa_emissora",
        "status",
        "local_calibracao",
    )

    search_fields = (
        "instrumento__codigo",
        "instrumento__descricao",
        "certificado_numero",
    )

    autocomplete_fields = ("instrumento",)

    readonly_fields = (
        "validade",
        "created_at",
        "marca",
        "modelo",
        "numero_serie",
        "local_instalacao",
        "ver_certificados_padroes",
        "ver_metodo",
    )

    fieldsets = (
        ("Instrumento", {
            "fields": ("instrumento",)
        }),

        ("Dados do Equipamento", {
            "fields": (
                "marca",
                "modelo",
                "numero_serie",
                "local_instalacao",
            )
        }),

        ("Dados da Calibração", {
            "fields": (
                "data_calibracao",
                "validade",
                "status",
                "resultado",
                "local_calibracao",
            )
        }),

        ("Condições Ambientais", {
            "fields": (
                "temperatura",
                "umidade",
            )
        }),

        ("Metodologia", {
            "fields": (
                "metodo",
                "padroes",
                "ver_metodo",
                "ver_certificados_padroes",
            )
        }),

        ("Execução", {
            "fields": (
                "equipamentos_auxiliares",
            )
        }),

        ("Certificado", {
            "fields": (
                "certificado_numero",
                "certificado_arquivo",
                "empresa_emissora",
            )
        }),

        ("Observações", {
            "fields": ("observacoes",)
        }),
    )

    def marca(self, obj):
        return getattr(obj.instrumento, "marca", "")

    def modelo(self, obj):
        return getattr(obj.instrumento, "modelo", "")

    def numero_serie(self, obj):
        return getattr(obj.instrumento, "numero_serie", "")

    def local_instalacao(self, obj):
        return getattr(obj.instrumento, "local_instalacao", "")

    def ver_metodo(self, obj):
        if obj.instrumento and obj.instrumento.metodo_calibracao:
            return obj.instrumento.metodo_calibracao.titulo
        return "—"

    def ver_certificados_padroes(self, obj):
        if not obj.instrumento:
            return "?"

        links = [
            (p.certificado.url, p.codigo)
            for p in obj.instrumento.padroes.all()
            if p.certificado
        ]
        if not links:
            return "?"
        return format_html_join(
            mark_safe("<br>"),
            '<a href="{}" target="_blank">{}</a>',
            links,
        )
    
    def ver_metodo(self, obj):

        if obj.instrumento and obj.instrumento.metodo_calibracao:

            metodo = obj.instrumento.metodo_calibracao

            if metodo.arquivo:

                return format_html(
                    '<a href="{}" target="_blank">{} - {}</a>',
                    metodo.arquivo.url,
                    metodo.codigo,
                    metodo.titulo
                )

        return "—"

    ver_metodo.short_description = "Ver método"

    def ver_anexo(self, obj):
        if obj.certificado_arquivo:
            return format_html(
                "<a href='{}' target='_blank'>Abrir</a>",
                obj.certificado_arquivo.url
            )
        return "—"

    ver_anexo.short_description = "Certificado"
    inlines = [CalibracaoAnexoInline]


@admin.register(CalibracaoTurbidez)
class CalibracaoTurbidezAdmin(admin.ModelAdmin):

    change_form_template = "admin/calibracao/calibracaoturbidez/change_form.html"

    list_display = (
        "numero_certificado",
        "instrumento",
        "cliente",
        "data_calibracao",
        "data_emissao",
        "pdf_certificado",
    )

    list_filter = ("data_calibracao", "data_emissao", "cliente")

    search_fields = (
        "numero_certificado",
        "instrumento__codigo",
        "instrumento__descricao",
        "cliente__razao_social",
        "ordem_servico",
    )

    autocomplete_fields = (
        "instrumento",
        "cliente",
        "procedimento_documento",
        "responsavel_tecnico_ref",
        "tecnico_executante_ref",
    )

    readonly_fields = ("numero_certificado", "procedimento_numero", "procedimento_revisao")

    fieldsets = (
        ("Planilha - Informações Gerais", {
            "fields": (
                "numero_certificado",
                "ordem_servico",
                "data_calibracao",
                "data_emissao",
                "revisao",
                "instrumento",
                "cliente",
                "contratante",
                "endereco_contratante",
                "endereco_cliente",
                "local_calibracao",
            )
        }),
        ("Planilha - Equipamento", {
            "fields": (
                "equipamento_calibrado",
                "numero_identificacao",
                "capacidade_total",
                "faixa_calibrada",
                "menor_resolucao",
                "unidade_leitura",
            )
        }),
        ("Planilha - Sonda Multiparâmetro", {
            "fields": (
                "sonda_multiparametro",
                "serie_sonda_optica_turbidez",
                "serie_cabo_multi",
                "id_sonda_multiparametro",
                "id_sonda_optica_turbidez",
                "id_cabo_multi",
            )
        }),
        ("Certificado - Procedimento e Ambiente", {
            "fields": (
                "procedimento_documento",
                "procedimento_numero",
                "procedimento_revisao",
                "temperatura_inicial",
                "temperatura_final",
                "umidade_inicial",
                "umidade_final",
                "ajuste_efetuado",
            )
        }),
        ("Certificado - Responsáveis e Observações", {
            "fields": (
                "responsavel_tecnico_ref",
                "tecnico_executante_ref",
                "funcao_signatario",
                "resultado_final",
                "observacoes_certificado",
            )
        }),
    )

    inlines = [
        TurbidezPadraoInline,
        TurbidezVerificacaoInline,
        TurbidezCalibracaoInline,
        TurbidezIncertezaInline,
    ]

    class Media:
        js = ("js/calibracao_turbidez.js",)

    def get_form(self, request, obj=None, change=False, **kwargs):
        form = super().get_form(request, obj=obj, change=change, **kwargs)
        instrumento_field = form.base_fields.get("instrumento")
        if instrumento_field:
            instrumento_url = reverse(
                "admin:calibracao_calibracaoturbidez_instrumento_dados",
                args=["00000000-0000-0000-0000-000000000000"],
            )
            instrumento_field.widget.attrs["data-instrumento-dados-url"] = instrumento_url.replace(
                "00000000-0000-0000-0000-000000000000",
                "__instrumento__",
            )
        return form

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        responsavel = ResponsavelCertificado.objects.filter(nome="Diego Henrique Alves Saldanha").first()
        if responsavel:
            initial.setdefault("responsavel_tecnico_ref", responsavel.pk)
            initial.setdefault("tecnico_executante_ref", responsavel.pk)
            initial.setdefault("funcao_signatario", responsavel.cargo)
        return initial

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "procedimento_documento":
            kwargs["queryset"] = Documento.objects.filter(
                tipo__in=("procedimento", "instrucao", "metodo"),
                status="vigente",
            ).order_by("codigo", "titulo")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "instrumento/<uuid:instrumento_id>/dados/",
                self.admin_site.admin_view(self.instrumento_dados_view),
                name="calibracao_calibracaoturbidez_instrumento_dados",
            ),
            path(
                "<uuid:pk>/pdf/",
                self.admin_site.admin_view(self.pdf_view),
                name="calibracao_calibracaoturbidez_pdf",
            ),
        ]
        return custom_urls + urls

    def pdf_view(self, request, pk):
        from .views import pdf_calibracao_turbidez

        return pdf_calibracao_turbidez(request, pk)

    def instrumento_dados_view(self, request, instrumento_id):
        instrumento = get_object_or_404(Instrumento.objects.select_related("cliente"), pk=instrumento_id)
        cliente = instrumento.cliente

        try:
            tecnico = instrumento.tecnico
        except InstrumentoTecnico.DoesNotExist:
            tecnico = None

        endereco_partes = [
            getattr(cliente, "endereco", ""),
            getattr(cliente, "numero", ""),
            getattr(cliente, "bairro", ""),
            getattr(cliente, "cidade", ""),
            getattr(cliente, "uf", ""),
        ]
        endereco_cliente = ", ".join([parte for parte in endereco_partes if parte])
        local_calibracao = CalibracaoTurbidez._normalizar_local_calibracao(instrumento.local_instalacao)

        return JsonResponse({
            "cliente": {
                "id": str(cliente.pk),
                "text": str(cliente),
                "razao_social": cliente.razao_social,
            },
            "contratante": cliente.razao_social or "",
            "endereco_contratante": endereco_cliente,
            "endereco_cliente": endereco_cliente,
            "local_calibracao": local_calibracao,
            "equipamento_calibrado": instrumento.descricao or "",
            "numero_identificacao": instrumento.codigo or "",
            "capacidade_total": (tecnico.capacidade_total if tecnico and tecnico.capacidade_total else instrumento.modelo or ""),
            "faixa_calibrada": tecnico.faixa_medicao if tecnico else "",
            "menor_resolucao": str(tecnico.menor_resolucao) if tecnico and tecnico.menor_resolucao is not None else "",
            "unidade_leitura": tecnico.unidade if tecnico else "",
            "numero_serie": instrumento.numero_serie or "",
            "marca": instrumento.marca or "",
            "modelo": instrumento.modelo or "",
        })

    def pdf_certificado(self, obj):
        return format_html(
            "<a class='button' href='{}' target='_blank'>Gerar PDF</a>",
            reverse("admin:calibracao_calibracaoturbidez_pdf", args=[obj.pk]),
        )

    pdf_certificado.short_description = "Certificado"


@admin.register(CalibracaoColorimetro)
class CalibracaoColorimetroAdmin(admin.ModelAdmin):

    change_form_template = "admin/calibracao/calibracaocolorimetro/change_form.html"

    list_display = (
        "numero_certificado",
        "tipo_aplicacao",
        "instrumento",
        "cliente",
        "data_calibracao",
        "data_emissao",
        "pdf_certificado",
    )

    list_filter = ("tipo_aplicacao", "data_calibracao", "data_emissao", "cliente")

    search_fields = (
        "numero_certificado",
        "instrumento__codigo",
        "instrumento__descricao",
        "cliente__razao_social",
        "ordem_servico",
    )

    autocomplete_fields = (
        "instrumento",
        "cliente",
        "procedimento_documento",
        "responsavel_tecnico_ref",
        "tecnico_executante_ref",
    )

    readonly_fields = ("numero_certificado", "procedimento_numero", "procedimento_revisao")

    fieldsets = (
        ("Planilha - Informações Gerais", {
            "fields": (
                "numero_certificado",
                "ordem_servico",
                "data_calibracao",
                "data_emissao",
                "revisao",
                "instrumento",
                "cliente",
                "contratante",
                "endereco_contratante",
                "endereco_cliente",
                "local_calibracao",
            )
        }),
        ("Planilha - Aplicação", {
            "fields": (
                "tipo_aplicacao",
                "unidade_leitura",
            )
        }),
        ("Planilha - Equipamento", {
            "fields": (
                "equipamento_calibrado",
                "numero_identificacao",
                "capacidade_total",
                "faixa_calibrada",
                "menor_resolucao",
            )
        }),
        ("Certificado - Procedimento e Ambiente", {
            "fields": (
                "procedimento_documento",
                "procedimento_numero",
                "procedimento_revisao",
                "temperatura_inicial",
                "temperatura_final",
                "umidade_inicial",
                "umidade_final",
                "ajuste_efetuado",
            )
        }),
        ("Certificado - Responsáveis e Observações", {
            "fields": (
                "responsavel_tecnico_ref",
                "tecnico_executante_ref",
                "funcao_signatario",
                "resultado_final_status",
                "resultado_final",
                "observacoes_certificado",
            )
        }),
    )

    inlines = [
        ColorimetroPadraoInline,
        ColorimetroVerificacaoInline,
        ColorimetroCalibracaoInline,
        ColorimetroIncertezaInline,
    ]

    class Media:
        js = ("js/calibracao_turbidez.js",)

    def get_form(self, request, obj=None, change=False, **kwargs):
        form = super().get_form(request, obj=obj, change=change, **kwargs)
        instrumento_field = form.base_fields.get("instrumento")
        tipo_field = form.base_fields.get("tipo_aplicacao")
        if instrumento_field:
            instrumento_url = reverse(
                "admin:calibracao_calibracaocolorimetro_instrumento_dados",
                args=["00000000-0000-0000-0000-000000000000"],
            )
            instrumento_field.widget.attrs["data-instrumento-dados-url"] = instrumento_url.replace(
                "00000000-0000-0000-0000-000000000000",
                "__instrumento__",
            )
        if tipo_field:
            metodo_url = reverse(
                "admin:calibracao_calibracaocolorimetro_metodo_dados",
                args=["__tipo__"],
            )
            tipo_field.widget.attrs["data-metodo-dados-url"] = metodo_url
        return form

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        responsavel = ResponsavelCertificado.objects.filter(nome="Diego Henrique Alves Saldanha").first()
        if responsavel:
            initial.setdefault("responsavel_tecnico_ref", responsavel.pk)
            initial.setdefault("tecnico_executante_ref", responsavel.pk)
            initial.setdefault("funcao_signatario", responsavel.cargo)
        return initial

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "procedimento_documento":
            kwargs["queryset"] = Documento.objects.filter(
                tipo__in=("procedimento", "instrucao", "metodo"),
                status="vigente",
            ).order_by("codigo", "titulo")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "instrumento/<uuid:instrumento_id>/dados/",
                self.admin_site.admin_view(self.instrumento_dados_view),
                name="calibracao_calibracaocolorimetro_instrumento_dados",
            ),
            path(
                "metodo/<str:tipo_aplicacao>/dados/",
                self.admin_site.admin_view(self.metodo_dados_view),
                name="calibracao_calibracaocolorimetro_metodo_dados",
            ),
            path(
                "<uuid:pk>/pdf/",
                self.admin_site.admin_view(self.pdf_view),
                name="calibracao_calibracaocolorimetro_pdf",
            ),
        ]
        return custom_urls + urls

    def instrumento_dados_view(self, request, instrumento_id):
        instrumento = get_object_or_404(Instrumento.objects.select_related("cliente"), pk=instrumento_id)
        cliente = instrumento.cliente
        try:
            tecnico = instrumento.tecnico
        except InstrumentoTecnico.DoesNotExist:
            tecnico = None

        endereco_partes = [
            getattr(cliente, "endereco", ""),
            getattr(cliente, "numero", ""),
            getattr(cliente, "bairro", ""),
            getattr(cliente, "cidade", ""),
            getattr(cliente, "uf", ""),
        ]
        endereco_cliente = ", ".join([parte for parte in endereco_partes if parte])
        local_calibracao = CalibracaoColorimetro._normalizar_local_calibracao(instrumento.local_instalacao)

        return JsonResponse({
            "cliente": {
                "id": str(cliente.pk),
                "text": str(cliente),
                "razao_social": cliente.razao_social,
            },
            "contratante": cliente.razao_social or "",
            "endereco_contratante": endereco_cliente,
            "endereco_cliente": endereco_cliente,
            "local_calibracao": local_calibracao,
            "equipamento_calibrado": instrumento.descricao or "Colorímetro",
            "numero_identificacao": instrumento.codigo or "",
            "capacidade_total": (tecnico.capacidade_total if tecnico and tecnico.capacidade_total else instrumento.modelo or ""),
            "faixa_calibrada": tecnico.faixa_medicao if tecnico else "",
            "menor_resolucao": str(tecnico.menor_resolucao) if tecnico and tecnico.menor_resolucao is not None else "",
            "unidade_leitura": tecnico.unidade if tecnico else "",
            "numero_serie": instrumento.numero_serie or "",
            "marca": instrumento.marca or "",
            "modelo": instrumento.modelo or "",
        })

    def pdf_view(self, request, pk):
        from .views import pdf_calibracao_colorimetro
        return pdf_calibracao_colorimetro(request, pk)

    def metodo_dados_view(self, request, tipo_aplicacao):
        calibracao_dummy = CalibracaoColorimetro(tipo_aplicacao=tipo_aplicacao)
        documento = calibracao_dummy._obter_documento_metodo_padrao()
        if not documento:
            return JsonResponse({})
        return JsonResponse({
            "documento": {
                "id": str(documento.pk),
                "text": str(documento),
            },
            "codigo": documento.codigo or "",
            "revisao": documento.revisao or "",
            "arquivo_url": documento.arquivo.url if documento.arquivo else "",
        })

    def pdf_certificado(self, obj):
        return format_html(
            "<a class='button' href='{}' target='_blank'>Gerar PDF</a>",
            reverse("admin:calibracao_calibracaocolorimetro_pdf", args=[obj.pk]),
        )

    pdf_certificado.short_description = "Certificado"

    def arquivo_metodo_pdf(self, obj):
        documento = getattr(obj, "procedimento_documento", None)
        if documento and documento.arquivo:
            return format_html("<a href='{}' target='_blank'>Abrir PDF do método</a>", documento.arquivo.url)
        return "—"

    arquivo_metodo_pdf.short_description = "PDF do método"

# =========================
# ORDEM DE SERVIÇO
# =========================

@admin.register(OrdemServico)
class OrdemServicoAdmin(admin.ModelAdmin):

    list_display = (
        "numero",
        "cliente",
        "proposta",
        "status",
        "data_abertura",
        "data_conclusao",
        "total_equipamentos",
        "ver_anexo",
    )

    list_filter = (
        "status",
        "data_abertura",
        "data_conclusao",
    )

    search_fields = (
        "numero",
        "cliente__razao_social",
        "cliente__cnpj",
        "proposta__codigo",
        "instrumentos__codigo",
        "instrumentos__descricao",
    )

    autocomplete_fields = (
        "cliente",
        "proposta",
        "instrumentos",
    )

    filter_horizontal = ("instrumentos",)

    readonly_fields = (
        "data_abertura",
        "ver_anexo",
    )

    fieldsets = (
        ("Dados da Ordem de Serviço", {
            "fields": (
                "numero",
                "proposta",
                "cliente",
                "status",
                "data_abertura",
                "data_conclusao",
            )
        }),
        ("Equipamentos", {
            "description": "Selecione apenas os equipamentos relacionados a esta OS.",
            "fields": (
                "instrumentos",
            )
        }),
        ("Anexo", {
            "fields": (
                "anexo",
                "ver_anexo",
            )
        }),
    )

    def ver_anexo(self, obj):
        if obj.anexo:
            return format_html(
                "<a href='{}' target='_blank'>Abrir</a>",
                obj.anexo.url
            )
        return "—"

    ver_anexo.short_description = "Anexo"

    def total_equipamentos(self, obj):
        return obj.instrumentos.count()

    total_equipamentos.short_description = "Equipamentos"


# =========================
# PADRÃO
# =========================

@admin.register(Padrao)
class PadraoAdmin(admin.ModelAdmin):

    list_display = (
        "codigo",
        "descricao",
        "numero_certificado",
        "laboratorio_emitente",
        "data_calibracao",
        "status",
        "vencimento",
        "ver_certificado",
    )

    search_fields = (
        "codigo",
        "descricao",
        "numero_certificado",
        "laboratorio_emitente",
    )

    list_filter = (
        "status",
    )

    ordering = ("codigo",)

    fieldsets = (
        ("Identificação", {
            "fields": (
                "codigo",
                "descricao",
                "status",
                "vencimento",
                "certificado",
            )
        }),
        ("Dados metrológicos", {
            "fields": (
                "numero_certificado",
                "laboratorio_emitente",
                "data_calibracao",
                "valor_nominal",
                "resolucao",
                "incerteza",
                "fator_k",
                "graus_liberdade",
                "unidade",
            )
        }),
    )

    def ver_certificado(self, obj):
        if obj.certificado:
            return format_html(
                "<a href='{}' target='_blank'>Abrir</a>",
                obj.certificado.url
            )
        return "—"

    ver_certificado.short_description = "Certificado"


@admin.register(CalibracaoPressao)
class CalibracaoPressaoAdmin(admin.ModelAdmin):
    change_form_template = "admin/calibracao/calibracaopressao/change_form.html"
    form = CalibracaoPressaoAdminForm

    list_display = (
        "numero_certificado",
        "tipo_instrumento",
        "instrumento",
        "cliente",
        "data_calibracao",
        "data_emissao",
        "pdf_certificado",
    )
    list_filter = ("tipo_instrumento", "tipo_indicacao", "data_calibracao", "data_emissao", "cliente")
    search_fields = (
        "numero_certificado",
        "instrumento__codigo",
        "instrumento__descricao",
        "cliente__razao_social",
        "ordem_servico",
    )
    autocomplete_fields = (
        "instrumento",
        "cliente",
        "procedimento_documento",
        "responsavel_tecnico_ref",
        "tecnico_executante_ref",
    )
    readonly_fields = ("numero_certificado", "procedimento_numero", "procedimento_revisao", "valor_por_divisao")
    fieldsets = (
        ("Planilha - InformaÃ§Ãµes Gerais", {
            "fields": (
                "numero_certificado",
                "ordem_servico",
                "data_calibracao",
                "data_emissao",
                "revisao",
                "instrumento",
                "cliente",
                "contratante",
                "endereco_contratante",
                "endereco_cliente",
                "local_calibracao",
            )
        }),
        ("Planilha - Equipamento", {
            "fields": (
                "tipo_instrumento",
                "tipo_indicacao",
                "referencia_calculo",
                "equipamento_calibrado",
                "numero_identificacao",
                "marca",
                "modelo",
                "numero_serie",
                "faixa_indicacao",
                "faixa_calibrada",
                "capacidade_total",
                "menor_resolucao",
                "classe_declarada",
                "unidade_indicacao",
                "unidade_padrao",
                "divisao_escala",
                "valor_por_divisao",
            )
        }),
        ("Certificado - MÃ©todo e Ambiente", {
            "fields": (
                "procedimento_documento",
                "procedimento_numero",
                "procedimento_revisao",
                "temperatura_inicial",
                "temperatura_final",
                "umidade_inicial",
                "umidade_final",
                "ajuste_efetuado",
            )
        }),
        ("Certificado - ResponsÃ¡veis e ObservaÃ§Ãµes", {
            "fields": (
                "responsavel_tecnico_ref",
                "tecnico_executante_ref",
                "funcao_signatario",
                "resultado_final_status",
                "resultado_final",
                "observacoes_certificado",
            )
        }),
    )
    inlines = [PressaoPadraoInline, PressaoCalibracaoInline, PressaoIncertezaInline]

    class Media:
        js = ("js/calibracao_pressao.js",)

    def get_form(self, request, obj=None, change=False, **kwargs):
        form = super().get_form(request, obj=obj, change=change, **kwargs)
        instrumento_field = form.base_fields.get("instrumento")
        if instrumento_field:
            instrumento_url = reverse(
                "admin:calibracao_calibracaopressao_instrumento_dados",
                args=["00000000-0000-0000-0000-000000000000"],
            )
            instrumento_field.widget.attrs["data-instrumento-dados-url"] = instrumento_url.replace(
                "00000000-0000-0000-0000-000000000000",
                "__instrumento__",
            )
        return form

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        responsavel = ResponsavelCertificado.objects.filter(nome="Diego Henrique Alves Saldanha").first()
        if responsavel:
            initial.setdefault("responsavel_tecnico_ref", responsavel.pk)
            initial.setdefault("tecnico_executante_ref", responsavel.pk)
            initial.setdefault("funcao_signatario", responsavel.cargo)
        return initial

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "procedimento_documento":
            kwargs["queryset"] = Documento.objects.filter(
                tipo__in=("procedimento", "instrucao", "metodo"),
                status="vigente",
            ).order_by("codigo", "titulo")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "instrumento/<uuid:instrumento_id>/dados/",
                self.admin_site.admin_view(self.instrumento_dados_view),
                name="calibracao_calibracaopressao_instrumento_dados",
            ),
            path(
                "<uuid:pk>/pdf/",
                self.admin_site.admin_view(self.pdf_view),
                name="calibracao_calibracaopressao_pdf",
            ),
        ]
        return custom_urls + urls

    def instrumento_dados_view(self, request, instrumento_id):
        instrumento = get_object_or_404(Instrumento.objects.select_related("cliente"), pk=instrumento_id)
        cliente = instrumento.cliente
        try:
            tecnico = instrumento.tecnico
        except InstrumentoTecnico.DoesNotExist:
            tecnico = None

        endereco_partes = [
            getattr(cliente, "endereco", ""),
            getattr(cliente, "numero", ""),
            getattr(cliente, "bairro", ""),
            getattr(cliente, "cidade", ""),
            getattr(cliente, "uf", ""),
        ]
        endereco_cliente = ", ".join([parte for parte in endereco_partes if parte])
        local_calibracao = CalibracaoPressao._normalizar_local_calibracao(instrumento.local_instalacao)

        return JsonResponse({
            "cliente": {
                "id": str(cliente.pk),
                "text": str(cliente),
                "razao_social": cliente.razao_social,
            },
            "contratante": cliente.razao_social or "",
            "endereco_contratante": endereco_cliente,
            "endereco_cliente": endereco_cliente,
            "local_calibracao": local_calibracao,
            "equipamento_calibrado": instrumento.descricao or "Instrumento de PressÃ£o",
            "numero_identificacao": instrumento.codigo or "",
            "capacidade_total": (tecnico.capacidade_total if tecnico and tecnico.capacidade_total else instrumento.modelo or ""),
            "faixa_calibrada": tecnico.faixa_medicao if tecnico else "",
            "faixa_indicacao": tecnico.faixa_medicao if tecnico else "",
            "menor_resolucao": str(tecnico.menor_resolucao) if tecnico and tecnico.menor_resolucao is not None else "",
            "unidade_leitura": tecnico.unidade if tecnico else "",
            "unidade_indicacao": tecnico.unidade if tecnico else "",
            "numero_serie": instrumento.numero_serie or "",
            "marca": instrumento.marca or "",
            "modelo": instrumento.modelo or "",
            "classe_declarada": tecnico.classe if tecnico else "",
        })

    def pdf_view(self, request, pk):
        from .views import pdf_calibracao_pressao
        return pdf_calibracao_pressao(request, pk)

    def pdf_certificado(self, obj):
        return format_html(
            "<a class='button' href='{}' target='_blank'>Gerar PDF</a>",
            reverse("admin:calibracao_calibracaopressao_pdf", args=[obj.pk]),
        )

    pdf_certificado.short_description = "Certificado"


class PHPadraoInline(admin.TabularInline):
    model = CalibracaoPHPadraoUtilizado
    extra = 1
    autocomplete_fields = ("padrao",)
    fields = (
        "tipo",
        "ordem",
        "padrao",
        "codigo",
        "descricao",
        "numero_certificado",
        "laboratorio_emitente",
        "data_calibracao",
        "validade",
        "resolucao",
        "incerteza",
        "fator_k",
        "graus_liberdade",
        "unidade",
        "valor_nominal",
    )
    verbose_name = "Padrao"
    verbose_name_plural = "Padroes"


class PHCalibracaoEletricaInline(admin.TabularInline):
    model = CalibracaoPHPonto
    extra = 3
    fields = (
        "ordem",
        "tipo",
        "valor_padrao_mv",
        "leitura_1",
        "leitura_2",
        "leitura_3",
        "media",
        "desvio_padrao",
        "erro",
        "ema",
        "criterio",
        "resultado",
    )
    readonly_fields = ("media", "desvio_padrao", "erro", "ema")
    verbose_name = "Ponto de calibracao eletrica"
    verbose_name_plural = "Calibracao eletrica"

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.filter(tipo__in=("eletrica_mv", "eletrica_ph"))


class PHCalibracaoMRCInline(admin.TabularInline):
    model = CalibracaoPHPonto
    extra = 5
    fields = (
        "ordem",
        "tipo",
        "valor_padrao_ph",
        "leitura_1",
        "leitura_2",
        "leitura_3",
        "media",
        "desvio_padrao",
        "erro",
        "ema",
        "criterio",
        "resultado",
    )
    readonly_fields = ("media", "desvio_padrao", "erro", "ema")
    verbose_name = "Ponto de calibracao da parte quimica"
    verbose_name_plural = "Calibracao com parte quimica"

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.filter(
            tipo__in=(
                "quimica_acida",
                "quimica_neutra",
                "quimica_basica",
                "verificacao_acida",
                "verificacao_basica",
                "mrc_acida",
                "mrc_neutra",
                "mrc_basica",
                "mrc_verificacao_acida",
                "mrc_verificacao_basica",
            )
        )


class PHIncertezaInline(admin.TabularInline):
    model = CalibracaoPHIncertezaPonto
    extra = 1
    fields = (
        "ordem",
        "repetibilidade",
        "resolucao_instrumento",
        "resolucao_padrao",
        "incerteza_padrao",
        "incerteza_curva",
        "incerteza_temperatura",
        "incerteza_constante_faraday",
        "incerteza_constante_gas",
        "incerteza_phx",
        "fator_k",
        "graus_liberdade",
        "incerteza_padrao_combinada",
        "incerteza_expandida",
    )
    readonly_fields = ("incerteza_padrao_combinada", "incerteza_expandida")
    verbose_name = "Ponto de incerteza"
    verbose_name_plural = "Incerteza"


@admin.register(CalibracaoPH)
class CalibracaoPHAdmin(admin.ModelAdmin):

    change_form_template = "admin/calibracao/calibracaoph/change_form.html"

    list_display = (
        "numero_certificado",
        "tipo_calibracao",
        "instrumento",
        "cliente",
        "data_calibracao",
        "data_emissao",
        "pdf_certificado",
    )

    list_filter = ("tipo_calibracao", "local_calibracao", "data_calibracao", "cliente")

    search_fields = (
        "numero_certificado",
        "instrumento__codigo",
        "instrumento__descricao",
        "cliente__razao_social",
        "ordem_servico",
    )

    autocomplete_fields = (
        "instrumento",
        "cliente",
        "procedimento_documento",
        "responsavel_tecnico_ref",
        "tecnico_executante_ref",
    )

    readonly_fields = ("numero_certificado", "procedimento_numero", "procedimento_revisao", "resultado_final_status", "resultado_final")

    fieldsets = (
        ("Planilha - Informacoes Gerais", {
            "fields": (
                "numero_certificado",
                "ordem_servico",
                "data_calibracao",
                "data_emissao",
                "revisao",
                "instrumento",
                "cliente",
                "contratante",
                "endereco_contratante",
                "endereco_cliente",
                "local_calibracao",
                "tipo_calibracao",
                "tipo_indicacao",
            )
        }),
        ("Planilha - Equipamento", {
            "fields": (
                "equipamento_calibrado",
                "numero_identificacao",
                "marca",
                "modelo",
                "numero_serie",
                "capacidade_total",
                "faixa_calibrada",
                "menor_resolucao",
                "resolucao_mv",
                "resolucao_ph",
                "identificacao_eletrodo",
                "resolucao_termometro",
                "temperatura_referencia",
                "slope_indicado",
                "id_sensor_temperatura",
                "unidade_leitura",
                "tipo_sensor_temperatura",
            )
        }),
        ("Certificado - Procedimento e Ambiente", {
            "fields": (
                "procedimento_documento",
                "procedimento_numero",
                "procedimento_revisao",
                "temperatura_inicial",
                "temperatura_final",
                "umidade_inicial",
                "umidade_final",
                "ajuste_efetuado",
            )
        }),
        ("Certificado - Responsaveis e Observacoes", {
            "fields": (
                "responsavel_tecnico_ref",
                "tecnico_executante_ref",
                "funcao_signatario",
                "signatario_autorizado",
                "resultado_final_status",
                "resultado_final",
                "observacoes_certificado",
            )
        }),
    )

    inlines = [
        PHPadraoInline,
        PHCalibracaoEletricaInline,
        PHCalibracaoMRCInline,
        PHIncertezaInline,
    ]

    class Media:
        js = ("js/calibracao_ph.js",)

    def get_form(self, request, obj=None, change=False, **kwargs):
        form = super().get_form(request, obj=obj, change=change, **kwargs)
        instrumento_field = form.base_fields.get("instrumento")
        if instrumento_field:
            instrumento_url = reverse(
                "admin:calibracao_calibracaoph_instrumento_dados",
                args=["00000000-0000-0000-0000-000000000000"],
            )
            instrumento_field.widget.attrs["data-instrumento-dados-url"] = instrumento_url.replace(
                "00000000-0000-0000-0000-000000000000",
                "__instrumento__",
            )
        return form

    def get_changeform_initial_data(self, request):
        initial = super().get_changeform_initial_data(request)
        responsavel = ResponsavelCertificado.objects.filter(nome="Diego Henrique Alves Saldanha").first()
        if responsavel:
            initial.setdefault("responsavel_tecnico_ref", responsavel.pk)
            initial.setdefault("tecnico_executante_ref", responsavel.pk)
            initial.setdefault("funcao_signatario", responsavel.cargo)
        return initial

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "procedimento_documento":
            kwargs["queryset"] = Documento.objects.filter(
                tipo__in=("procedimento", "instrucao", "metodo"),
                status="vigente",
            ).order_by("codigo", "titulo")
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "instrumento/<uuid:instrumento_id>/dados/",
                self.admin_site.admin_view(self.instrumento_dados_view),
                name="calibracao_calibracaoph_instrumento_dados",
            ),
            path(
                "<int:pk>/pdf/",
                self.admin_site.admin_view(self.pdf_view),
                name="calibracao_calibracaoph_pdf",
            ),
        ]
        return custom_urls + urls

    def instrumento_dados_view(self, request, instrumento_id):
        instrumento = get_object_or_404(Instrumento.objects.select_related("cliente"), pk=instrumento_id)
        cliente = instrumento.cliente
        try:
            tecnico = instrumento.tecnico
        except InstrumentoTecnico.DoesNotExist:
            tecnico = None

        return JsonResponse({
            "cliente": {
                "id": str(cliente.pk),
                "text": str(cliente),
                "razao_social": cliente.razao_social,
            },
            "contratante": cliente.razao_social or "",
            "endereco_contratante": _join_address(cliente),
            "endereco_cliente": _join_address(cliente),
            "local_calibracao": CalibracaoPH._normalizar_local_calibracao(instrumento.local_instalacao),
            "equipamento_calibrado": instrumento.descricao or "Medidor de pH",
            "numero_identificacao": instrumento.codigo or "",
            "marca": instrumento.marca or "",
            "modelo": instrumento.modelo or "",
            "numero_serie": instrumento.numero_serie or "",
            "capacidade_total": (tecnico.capacidade_total if tecnico and tecnico.capacidade_total else instrumento.modelo or ""),
            "faixa_calibrada": tecnico.faixa_medicao if tecnico else "",
            "menor_resolucao": str(tecnico.menor_resolucao) if tecnico and tecnico.menor_resolucao is not None else "",
            "unidade_leitura": tecnico.unidade if tecnico else "",
            "tipo_indicacao": "digital" if tecnico and tecnico.classe else "",
            "resolucao_mv": str(tecnico.menor_resolucao) if tecnico and tecnico.menor_resolucao is not None else "",
            "resolucao_ph": str(tecnico.menor_resolucao) if tecnico and tecnico.menor_resolucao is not None else "",
            "identificacao_eletrodo": instrumento.codigo or "",
            "id_sensor_temperatura": "",
        })

    def pdf_view(self, request, pk):
        from .views import pdf_calibracao_ph

        return pdf_calibracao_ph(request, pk)

    def pdf_certificado(self, obj):
        return format_html(
            "<a class='button' href='{}' target='_blank'>Gerar PDF</a>",
            reverse("admin:calibracao_calibracaoph_pdf", args=[obj.pk]),
        )

    pdf_certificado.short_description = "Certificado"


@admin.register(CalibracaoCondutividade)
class CalibracaoCondutividadeAdmin(admin.ModelAdmin):
    list_display = (
        "numero_certificado",
        "cliente",
        "instrumento",
        "data_calibracao",
        "resultado_final",
        "status",
    )
    list_filter = ("status", "resultado_final", "local_calibracao")
    search_fields = (
        "numero_certificado",
        "cliente__razao_social",
        "instrumento__codigo",
        "instrumento__descricao",
    )
    autocomplete_fields = ("cliente", "instrumento")
    fieldsets = (
        ("Identificação", {
            "fields": (
                "cliente",
                "instrumento",
                "ordem_servico",
                "numero_certificado",
                "revisao",
            )
        }),
        ("Datas e status", {
            "fields": (
                "data_calibracao",
                "data_emissao",
                "local_calibracao",
                "status",
                "resultado_final",
            )
        }),
        ("Dados do equipamento", {
            "fields": (
                "contratante",
                "endereco_contratante",
                "endereco_cliente",
                "equipamento_calibrado",
                "numero_identificacao",
                "marca",
                "modelo",
                "numero_serie",
                "faixa_capacidade",
                "resolucao",
                "unidade_indicacao",
                "unidade_padrao",
                "unidade_leitura",
                "temperatura_referencia",
                "identificacao_celula",
                "constante_celula",
                "compensacao_temperatura",
            )
        }),
        ("Responsáveis", {
            "fields": (
                "tecnico_responsavel",
                "responsavel_conferencia",
                "signatario_autorizado",
                "funcao_signatario",
            )
        }),
        ("Ambiente e controle", {
            "fields": (
                "temperatura_ambiente",
                "umidade_ambiente",
                "observacoes",
                "snapshot_json",
                "pdf_arquivo",
            )
        }),
    )
